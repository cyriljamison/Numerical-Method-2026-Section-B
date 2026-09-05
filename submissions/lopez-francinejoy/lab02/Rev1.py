"""
create_cube_rev1.py — 3D Structural Frame Model (Rev. 1)

A RISA/STAAD-style 3D frame model of a 6 m x 6 m x 6 m cube. Built on the
original create_cube.py geometry and extended with Rev. 1 modelling tools:

    1.  Support conditions at the bottom nodes  (all four base nodes Pinned).
    2.  A per-member beta angle (rotation of the local 2-3 axes about the
        member axis).
    3.  Local member axes  (1 = axial, 2 = strong, 3 = weak) and the global
        axes.
    4.  Degrees of freedom (DOF) for every node: DX, DY, DZ, RX, RY, RZ.
    5.  Beam end DOFs — a "Pinned in X" end releases the Moment-Z (Mz).
    6.  A Pinned symbol drawn at pinned supports and pinned beam ends.
    7.  Updated structural diagram  (supports, hinges, local & global axes).
    8.  Updated styled Excel workbook  (cube_structure_rev1.xlsx).
    9.  Delivered as Rev. 1  (this file + rev1 outputs).

Global convention:  X, Z = lateral axes,  Y = vertical (up).
Local member convention:  axis 1 = axial (i -> j), axis 2 = strong,
                          axis 3 = weak (moment about it is Mz).

Run:  python create_cube_rev1.py
"""

# =============================================================================
# IMPORTS
# =============================================================================
import math
import os

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401  (registers 3D projection)

# openpyxl styling helpers (guarded so the script runs even if missing)
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAS_OPENPYXL = False


# =============================================================================
# MODEL DATA
# =============================================================================
L = 6.0  # cube side length (m)

NODES = {
    1: (0.0, 0.0, 0.0),
    2: (L,   0.0, 0.0),
    3: (L,   0.0, L),
    4: (0.0, 0.0, L),
    5: (0.0, L,   0.0),
    6: (L,   L,   0.0),
    7: (L,   L,   L),
    8: (0.0, L,   L),
}

# members: (name, start node i, end node j)
MEMBERS = [
    ("M1",  1, 2),
    ("M2",  2, 3),
    ("M3",  3, 4),
    ("M4",  4, 1),
    ("M5",  5, 6),
    ("M6",  6, 7),
    ("M7",  7, 8),
    ("M8",  8, 5),
    ("M9",  1, 5),   # vertical column
    ("M10", 2, 6),   # vertical column
    ("M11", 3, 7),   # vertical column
    ("M12", 4, 8),   # vertical column
]

GLOBAL_DOFS = ["DX", "DY", "DZ", "RX", "RY", "RZ"]

SUPPORT_TYPE = "Pinned"                # "for now, use pinned supports"
PINNED_SUPPORT_DOFS = ["DX", "DY", "DZ"]  # a pin restrains the translations

# Member-end release codes (RISA/STAAD style):
#   'R' = continuous (rigid) end — moment fully transmitted
#   'P' = Pinned-in-X: releases the end moment about the local Z axis (Mz)
RELEASES = {
    "M1":  ("P", "P"),   # base ring ends pinned (example)
    "M2":  ("R", "R"),
    "M3":  ("R", "R"),
    "M4":  ("P", "P"),
    "M5":  ("R", "R"),
    "M6":  ("R", "R"),
    "M7":  ("R", "R"),
    "M8":  ("R", "R"),
    "M9":  ("P", "R"),   # column base pinned
    "M10": ("P", "R"),
    "M11": ("P", "R"),
    "M12": ("P", "R"),
}

BETA_DEFAULT = 0.0   # default beta angle for every member (degrees)

OUT_XLSX = "cube_structure_rev1.xlsx"
OUT_PNG = "cube_structure_rev1.png"


# =============================================================================
# GEOMETRY / LOCAL AXES
# =============================================================================
def _norm(v):
    """Normalise a vector (returns it unchanged if it has zero length)."""
    v = np.asarray(v, dtype=float)
    n = np.linalg.norm(v)
    return v / n if n > 0.0 else v


def local_axes(ni, nj, beta_deg):
    """
    Compute a member's local axis set (1, 2, 3) from its end nodes.

    - axis 1 is the axial direction (start node -> end node)
    - axis 3 (weak) is built perpendicular to axis 1 (global Y preferred)
    - axis 2 (strong) completes the right-handed triad
    - beta rotates axes 2 & 3 together about the axial axis
    """
    xi, yi, zi = NODES[ni]
    xj, yj, zj = NODES[nj]
    d = np.array([xj - xi, yj - yi, zj - zi])
    length = float(np.linalg.norm(d))

    a1 = _norm(d)                              # axial (local X)
    ref = np.array([0.0, 1.0, 0.0])            # global Y preferred
    if abs(float(np.dot(a1, ref))) > 0.999:    # member already vertical
        ref = np.array([0.0, 0.0, 1.0])        # fall back to global Z

    a3 = _norm(np.cross(a1, ref))              # local weak
    a2 = _norm(np.cross(a3, a1))               # local strong

    beta = math.radians(beta_deg)              # rotate 2-3 about axis 1
    ct, st = math.cos(beta), math.sin(beta)
    a2r = a2 * ct + a3 * st
    a3r = -a2 * st + a3 * ct

    return {
        "Length": length,
        "dx": d[0], "dy": d[1], "dz": d[2],
        "a1": a1, "a2": a2r, "a3": a3r,
        "beta": beta_deg,
        "l": d[0] / length, "m": d[1] / length, "n": d[2] / length,
    }


def release_text(code):
    """Human-readable description of an end-release code."""
    if code == "P":
        return "Pinned-in-X -> Moment-Z (Mz) released"
    return "Rigid (no release)"


def _fmt_vec(v, nd=4):
    """Format a vector as a comma-separated string of nd-decimal numbers."""
    return ", ".join(f"{val:.{nd}f}" for val in v)


# =============================================================================
# 1. NODE DOFs AND SUPPORT CONDITIONS
# =============================================================================
def build_node_dof():
    dof_types = {"D": "Translational", "R": "Rotational"}
    dof_rows, node_summary, supports = [], [], []

    global_index = 0
    for nid in sorted(NODES):
        is_support = NODES[nid][1] == 0.0                 # base node (Y = 0)
        restrained = set(PINNED_SUPPORT_DOFS) if is_support else set()

        for label in GLOBAL_DOFS:
            active = label not in restrained
            if active:
                global_index += 1
            dof_rows.append({
                "Node": nid, "DOF": label,
                "DOF Type": dof_types[label[0]],
                "Active": 1 if active else 0,
                "Restrained": 0 if active else 1,
                "Global DOF #": global_index if active else "-",
            })

        active_labels = [lb for lb in GLOBAL_DOFS if lb not in restrained]
        node_summary.append({
            "Node": nid, "X (m)": NODES[nid][0], "Y (m)": NODES[nid][1],
            "Z (m)": NODES[nid][2],
            "Support": SUPPORT_TYPE if is_support else "None",
            "Restrained DOFs": ", ".join(sorted(restrained)) if restrained else "-",
            "Active DOFs": ", ".join(active_labels) if active_labels else "-",
        })
        if is_support:
            supports.append({
                "Node": nid, "Support Type": SUPPORT_TYPE,
                "Restrained DOFs": ", ".join(sorted(restrained)),
                "Free DOFs": ", ".join(active_labels) if active_labels else "-",
            })

    return (pd.DataFrame(dof_rows), pd.DataFrame(node_summary),
            pd.DataFrame(supports), global_index)


# =============================================================================
# 2. MEMBER LOCAL AXES AND END RELEASES
# =============================================================================
def build_members():
    inc_rows, axes_rows, release_rows = [], [], []

    for name, ni, nj in MEMBERS:
        ax = local_axes(ni, nj, BETA_DEFAULT)
        xi, yi, zi = NODES[ni]
        xj, yj, zj = NODES[nj]

        inc_rows.append({
            "Member": name, "i": ni, "j": nj,
            "Xi (m)": xi, "Yi (m)": yi, "Zi (m)": zi,
            "Xj (m)": xj, "Yj (m)": yj, "Zj (m)": zj,
            "dx (m)": ax["dx"], "dy (m)": ax["dy"], "dz (m)": ax["dz"],
            "Length (m)": round(ax["Length"], 4),
            "Beta (deg)": ax["beta"],
            "l": round(ax["l"], 4), "m": round(ax["m"], 4), "n": round(ax["n"], 4),
        })

        axes_rows.append({
            "Member": name, "i": ni, "j": nj, "Beta (deg)": ax["beta"],
            "Local 1 X": _fmt_vec(ax["a1"]),
            "Local 2 Y (strong)": _fmt_vec(ax["a2"]),
            "Local 3 Z (weak)": _fmt_vec(ax["a3"]),
            "Mz axis (local 3)": _fmt_vec(ax["a3"]),
        })

        rel_i, rel_j = RELEASES.get(name, ("R", "R"))
        release_rows.append({
            "Member": name,
            "End i release": rel_i, "End i DOF": release_text(rel_i),
            "End j release": rel_j, "End j DOF": release_text(rel_j),
            "Pinned?": "Yes" if ("P" in (rel_i, rel_j)) else "No",
        })

    return (pd.DataFrame(inc_rows), pd.DataFrame(axes_rows),
            pd.DataFrame(release_rows))


# =============================================================================
# 3. EXCEL OUTPUT
# =============================================================================
def _build_summary(total_active):
    """Return a one-row-per-item model summary DataFrame."""
    return pd.DataFrame([
        {"Item": "Model", "Value": "6m x 6m x 6m space frame (Rev. 1)"},
        {"Item": "Nodes", "Value": len(NODES)},
        {"Item": "Members", "Value": len(MEMBERS)},
        {"Item": "DOF per node", "Value": len(GLOBAL_DOFS)},
        {"Item": "Total DOFs (model)", "Value": len(NODES) * len(GLOBAL_DOFS)},
        {"Item": "Support type", "Value": SUPPORT_TYPE},
        {"Item": "Pinned supports", "Value": ", ".join(
            str(n) for n in sorted(NODES) if NODES[n][1] == 0.0)},
        {"Item": "Pinned support restraints",
         "Value": ", ".join(PINNED_SUPPORT_DOFS)},
        {"Item": "Active (free) DOFs", "Value": total_active},
        {"Item": "Restrained DOFs",
         "Value": len(NODES) * len(GLOBAL_DOFS) - total_active},
        {"Item": "Beta default (deg)", "Value": BETA_DEFAULT},
    ])


def write_excel(dof_df, node_df, support_df, inc_df, axes_df, release_df,
                total_active, filename):
    """Write every DataFrame to its own styled sheet in the workbook."""
    sheets = [
        ("Summary", _build_summary(total_active)),
        ("Nodes", node_df),
        ("Node DOFs", dof_df),
        ("Supports", support_df),
        ("Incidences", inc_df),
        ("Member Local Axes", axes_df),
        ("Member End Releases", release_df),
    ]

    descriptions = {
        "Summary": "Model summary and DOF counts for the Rev. 1 space frame.",
        "Nodes": "Node coordinates (global) and support / DOF summary per node.",
        "Node DOFs": "Per-node degrees of freedom (translations + rotations).",
        "Supports": "Pinned support conditions applied at the base nodes.",
        "Incidences": "Member incidences, geometry and direction cosines.",
        "Member Local Axes": "Local 1-2-3 axes per member (beta applied).",
        "Member End Releases": "End releases - 'P' releases the Moment-Z (Mz).",
    }

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name, index=False)
        if _HAS_OPENPYXL:
            _style_workbook(writer.book, descriptions)

    print(f"Excel workbook written: {os.path.abspath(filename)}")


def _style_workbook(wb, descriptions):
    """Apply headers, widths, freeze panes, borders and banding to all sheets."""
    header_fill = PatternFill("solid", fgColor="304C89")   # deep blue
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    band_fill = PatternFill("solid", fgColor="EDF2FA")     # light blue band
    title_font = Font(bold=True, size=14, color="304C89", name="Calibri")
    thin = Side(style="thin", color="C8CFD8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        # Title banner row on top of every sheet.
        ws.insert_rows(1)
        title_cell = ws.cell(row=1, column=1,
                             value=f"3D STRUCTURAL FRAME - {ws.title.upper()}")
        title_cell.font = title_font
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=max(1, ws.max_column))
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # One-line description beneath the title.
        desc = descriptions.get(ws.title, "")
        header_row = 2
        if desc:
            ws.insert_rows(2)
            dcell = ws.cell(row=2, column=1, value=desc)
            dcell.font = Font(italic=True, size=9, color="44546A", name="Calibri")
            ws.merge_cells(start_row=2, start_column=1,
                           end_row=2, end_column=max(1, ws.max_column))
            dcell.alignment = left
            ws.row_dimensions[2].height = 26
            header_row = 3

        # Style the header row.
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=header_row, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border

        # Freeze panes below the header so it stays visible while scrolling.
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Data rows: borders, alignment, alternating banding.
        for r in range(header_row + 1, ws.max_row + 1):
            band = (r - header_row) % 2 == 0
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=r, column=col)
                c.border = border
                c.alignment = left if col == 1 else center
                if band:
                    c.fill = band_fill

        # Column widths from content (capped).
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            longest = 0
            for r in range(header_row, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    longest = max(longest, len(str(v)))
            ws.column_dimensions[letter].width = min(max(longest + 4, 10), 55)

        # Autofilter on the header row.
        ws.auto_filter.ref = (
            f"{get_column_letter(1)}{header_row}:"
            f"{get_column_letter(ws.max_column)}{ws.max_row}"
        )
        ws.sheet_view.zoomScale = 100


# =============================================================================
# 4. STRUCTURAL DIAGRAM
# =============================================================================
def set_axes_equal(ax):
    """Give the 3D axes an equal (cubic) scale."""
    xl, yl, zl = ax.get_xlim3d(), ax.get_ylim3d(), ax.get_zlim3d()
    ranges = [abs(xl[1] - xl[0]), abs(yl[1] - yl[0]), abs(zl[1] - zl[0])]
    mids = [(xl[0] + xl[1]) / 2.0, (yl[0] + yl[1]) / 2.0, (zl[0] + zl[1]) / 2.0]
    r = 0.5 * max(ranges)
    ax.set_xlim3d([mids[0] - r, mids[0] + r])
    ax.set_ylim3d([mids[1] - r, mids[1] + r])
    ax.set_zlim3d([mids[2] - r, mids[2] + r])


def draw_pin_support(ax, x, y, z, s=0.45):
    """Draw a downward triangle (pin symbol) below a support node."""
    apex = (x, y - 1.6 * s, z)
    bl = (x - 0.9 * s, y - 0.4 * s, z)
    br = (x + 0.9 * s, y - 0.4 * s, z)
    ax.plot3D([apex[0], bl[0]], [apex[1], bl[1]], [apex[2], bl[2]],
              color="k", lw=1.6)
    ax.plot3D([bl[0], br[0]], [bl[1], br[1]], [bl[2], br[2]], color="k", lw=1.6)
    ax.plot3D([br[0], apex[0]], [br[1], apex[1]], [br[2], apex[2]],
              color="k", lw=1.6)


def draw_arrow(ax, o, v, color, label, s=0.7):
    """Draw a labelled 3D axis arrow starting at o and pointing along v."""
    tip = o + v * s
    ax.plot3D([o[0], tip[0]], [o[1], tip[1]], [o[2], tip[2]], color=color, lw=2)
    ax.scatter(*tip, color=color, s=20)
    ax.text(*(tip + np.array([0.05, 0.05, 0.05])), label,
            color=color, fontsize=9, fontweight="bold")


def draw_diagram(inc_df):  # noqa: ARG001  (inc_df kept for future use)
    """Draw the model: members, nodes, supports, hinges and axes."""
    from matplotlib.lines import Line2D

    fig = plt.figure(figsize=(11, 8))
    ax = fig.add_subplot(111, projection="3d")

    for name, ni, nj in MEMBERS:
        axl = local_axes(ni, nj, BETA_DEFAULT)
        xi, yi, zi = NODES[ni]
        xj, yj, zj = NODES[nj]
        ax.plot3D([xi, xj], [yi, yj], [zi, zj], color="tab:blue", lw=2.5)

        xm, ym, zm = (xi + xj) / 2.0, (yi + yj) / 2.0, (zi + zj) / 2.0
        ax.text(xm, ym + 0.15, zm, name, color="tab:green", fontsize=9,
                ha="center", va="center", fontweight="bold")

        mid = np.array([xm, ym, zm])          # local axes at member midpoint
        draw_arrow(ax, mid, axl["a1"], "crimson", "1", s=0.55)
        draw_arrow(ax, mid, axl["a2"], "darkorange", "2", s=0.45)
        draw_arrow(ax, mid, axl["a3"], "purple", "3", s=0.45)

        # Hinge (pin) symbols at released ends.
        rel_i, rel_j = RELEASES.get(name, ("R", "R"))
        for end, code in (("i", rel_i), ("j", rel_j)):
            if code == "P":
                node = ni if end == "i" else nj
                ax.scatter(*NODES[node], color="k", s=80, marker="o",
                           facecolors="none", edgecolors="k", linewidths=1.8,
                           zorder=5)

    # Nodes.
    xs = [c[0] for c in NODES.values()]
    ys = [c[1] for c in NODES.values()]
    zs = [c[2] for c in NODES.values()]
    ax.scatter(xs, ys, zs, color="red", s=40)
    for nid, (x, y, z) in NODES.items():
        ax.text(x, y, z, str(nid), color="black", fontsize=10,
                ha="right", va="bottom", fontweight="bold")

    # Pinned support triangles at the base nodes.
    for nid in sorted(NODES):
        if NODES[nid][1] == 0.0:
            draw_pin_support(ax, *NODES[nid])

    # Global axes at the origin.
    origin = np.array([0.0, 0.0, 0.0])
    draw_arrow(ax, origin, np.array([1.0, 0.0, 0.0]), "red", "Xg", s=1.4)
    draw_arrow(ax, origin, np.array([0.0, 1.0, 0.0]), "green", "Yg", s=1.4)
    draw_arrow(ax, origin, np.array([0.0, 0.0, 1.0]), "blue", "Zg", s=1.4)

    ax.set_xlabel("X (global lateral)")
    ax.set_ylabel("Y (global vertical)")
    ax.set_zlabel("Z (global lateral)")
    ax.set_title("6m x 6m x 6m Cube - Rev.1  |  Pinned supports, Local axes 1-2-3, "
                 "Hinges, Global axes Xg-Yg-Zg")
    set_axes_equal(ax)

    legend = [
        Line2D([0], [0], color="tab:blue", lw=2.5, label="Member"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor="red",
               markersize=6, label="Node"),
        Line2D([0], [0], marker="o", color="w", markerfacecolor="none",
               markeredgecolor="k", markersize=9,
               label="Pinned end (Moment-Z released)"),
        Line2D([0], [0], marker="2", color="k", markersize=10,
               label="Pinned support (DX,DY,DZ restrained)"),
        Line2D([0], [0], color="crimson", lw=2, label="Local axis 1 (axial)"),
        Line2D([0], [0], color="darkorange", lw=2, label="Local axis 2 (strong)"),
        Line2D([0], [0], color="purple", lw=2, label="Local axis 3 (weak) / Mz"),
    ]
    ax.legend(handles=legend, loc="upper left", fontsize=8, ncol=2)
    fig.tight_layout()
    return fig, ax


# =============================================================================
# 5. MAIN DRIVER
# =============================================================================
def _is_interactive_backend():
    backend = matplotlib.get_backend().lower()
    return backend not in ("agg", "pdf", "svg", "ps", "template")


def main():
    dof_df, node_df, support_df, total_active = build_node_dof()
    inc_df, axes_df, release_df = build_members()

    write_excel(dof_df, node_df, support_df, inc_df, axes_df, release_df,
                total_active, OUT_XLSX)

    fig, ax = draw_diagram(inc_df)
    fig.savefig(OUT_PNG, dpi=200)
    print(f"Structural diagram saved: {os.path.abspath(OUT_PNG)}")

    # Console summary.
    base_nodes = [n for n in sorted(NODES) if NODES[n][1] == 0.0]
    print("\n======== REV.1 MODEL SUMMARY ========")
    print(f"Nodes: {len(NODES)}   Members: {len(MEMBERS)}")
    print(f"Support: {SUPPORT_TYPE} at base nodes {base_nodes}")
    print(f"Restraints per pinned node: {PINNED_SUPPORT_DOFS}")
    print(f"Total DOFs: {len(NODES) * len(GLOBAL_DOFS)}   "
          f"Active: {total_active}   Restrained: "
          f"{len(NODES) * len(GLOBAL_DOFS) - total_active}")
    print("Member end releases (P = Pinned-in-X -> Moment-Z released):")
    for name, rel in RELEASES.items():
        print(f"   {name}: i={rel[0]}, j={rel[1]}")
    print("=" * 32)

    # Show the interactive window unless the backend is headless.
    try:
        if _is_interactive_backend():
            plt.show()
    except Exception:
        pass


if __name__ == "__main__":
    import matplotlib
    main()
