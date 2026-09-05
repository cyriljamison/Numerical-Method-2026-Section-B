import math
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill

# Rev. 1 structural model: global X/Z are lateral directions and global Y is up.
members_data_rev1 = [
    # ID, Name, Start Node (i), End Node (j), Nominal Length, Description
    (1, "M1", 1, 2, 6.0, "Bottom Frame Front"),
    (2, "M2", 2, 3, 6.0, "Bottom Frame Right"),
    (3, "M3", 3, 4, 6.0, "Bottom Frame Back"),
    (4, "M4", 4, 1, 6.0, "Bottom Frame Left"),
    (5, "M5", 5, 6, 6.0, "Top Frame Front"),
    (6, "M6", 6, 7, 6.0, "Top Frame Right"),
    (7, "M7", 7, 8, 6.0, "Top Frame Back"),
    (8, "M8", 8, 5, 6.0, "Top Frame Left"),
    (9, "M9", 1, 5, 6.0, "Vertical Column Front-Left"),
    (10, "M10", 2, 6, 6.0, "Vertical Column Front-Right"),
    (11, "M11", 3, 7, 6.0, "Vertical Column Back-Right"),
    (12, "M12", 4, 8, 6.0, "Vertical Column Back-Left"),
]

nodes_data = [
    (1, 0.0, 0.0, 0.0, "Bottom-Left-Front (Origin)"),
    (2, 6.0, 0.0, 0.0, "Bottom-Right-Front"),
    (3, 6.0, 0.0, 6.0, "Bottom-Right-Back"),
    (4, 0.0, 0.0, 6.0, "Bottom-Left-Back"),
    (5, 0.0, 6.0, 0.0, "Top-Left-Front"),
    (6, 6.0, 6.0, 0.0, "Top-Right-Front"),
    (7, 6.0, 6.0, 6.0, "Top-Right-Back"),
    (8, 0.0, 6.0, 6.0, "Top-Left-Back")
]

DOF_LABELS = ("UX", "UY", "UZ", "RX", "RY", "RZ")
BOTTOM_NODES = {1, 2, 3, 4}
PINNED_MEMBERS = {"M1", "M3", "M5", "M7"}


def dof_numbers(node_id):
    first = (node_id - 1) * 6 + 1
    return {label: first + index for index, label in enumerate(DOF_LABELS)}


def member_axes(start, end, beta_degrees):
    direction = [end[i] - start[i] for i in range(3)]
    length = math.sqrt(sum(value * value for value in direction))
    local_x = [value / length for value in direction]
    reference = [0.0, 1.0, 0.0] if abs(local_x[1]) < 0.9 else [1.0, 0.0, 0.0]
    projection = sum(local_x[i] * reference[i] for i in range(3))
    local_y = [reference[i] - projection * local_x[i] for i in range(3)]
    local_y_length = math.sqrt(sum(value * value for value in local_y))
    local_y = [value / local_y_length for value in local_y]
    local_z = [
        local_x[1] * local_y[2] - local_x[2] * local_y[1],
        local_x[2] * local_y[0] - local_x[0] * local_y[2],
        local_x[0] * local_y[1] - local_x[1] * local_y[0],
    ]
    beta = math.radians(beta_degrees)
    beta_y = [local_y[i] * math.cos(beta) + local_z[i] * math.sin(beta) for i in range(3)]
    beta_z = [-local_y[i] * math.sin(beta) + local_z[i] * math.cos(beta) for i in range(3)]
    return length, local_x, beta_y, beta_z


def format_vector(vector):
    return "(" + ", ".join(f"{value:.3f}" for value in vector) + ")"


def draw_arrow(axis, start, vector, color, label, fontsize=9, linewidth=2):
    axis.quiver(*start, *vector, color=color, linewidth=linewidth, arrow_length_ratio=0.18)
    axis.text(start[0] + vector[0], start[1] + vector[1], start[2] + vector[2], label,
              color=color, weight="bold", fontsize=fontsize)


def plot_local_axes(axis, origin, local_x, local_y, local_z, member_name, scale=0.72):
    """Plot each member's beta-rotated local triad in plot coordinates."""
    plot_vectors = (
        (local_x[0], local_x[2], local_x[1]),
        (local_y[0], local_y[2], local_y[1]),
        (local_z[0], local_z[2], local_z[1]),
    )
    colors = ("#dc2626", "#16a34a", "#7c3aed")
    labels = (f"x{member_name}", f"y{member_name}", f"z{member_name}")
    for vector, color, label in zip(plot_vectors, colors, labels):
        scaled = tuple(value * scale for value in vector)
        axis.quiver(*origin, *scaled, color=color, linewidth=0.9, arrow_length_ratio=0.22, alpha=0.9)
        axis.text(origin[0] + scaled[0], origin[1] + scaled[1], origin[2] + scaled[2], label,
                  color=color, fontsize=5.5)


def create_structural_diagram(output_path):
    coordinates = {node[0]: node[1:4] for node in nodes_data}
    figure = plt.figure(figsize=(14, 8))
    axis = figure.add_subplot(111, projection="3d")
    axis.set_position([0.04, 0.08, 0.63, 0.82])
    for member in members_data_rev1:
        start = coordinates[member[2]]
        end = coordinates[member[3]]
        is_column = member[1] in {"M9", "M10", "M11", "M12"}
        color = "#14866d" if is_column else "#1554d1"
        axis.plot([start[0], end[0]], [start[2], end[2]], [start[1], end[1]], color=color, linewidth=2.5)
        midpoint = [(start[i] + end[i]) / 2 for i in range(3)]
        axis.text(midpoint[0], midpoint[2], midpoint[1] + 0.14, member[1], color=color, fontsize=8, weight="bold")
        beta = 90.0 if is_column else 0.0
        _, local_x, local_y, local_z = member_axes(start, end, beta)
        plot_local_axes(axis, (midpoint[0], midpoint[2], midpoint[1]), local_x, local_y, local_z, member[1])
        if member[1] in PINNED_MEMBERS:
            axis.scatter(midpoint[0], midpoint[2], midpoint[1], facecolors="white", edgecolors="#111827", s=58, linewidth=1.5)
            axis.text(midpoint[0] + 0.12, midpoint[2], midpoint[1], "[MZ]", color="#b91c1c", fontsize=7)
    for node_id, x, y, z, _ in nodes_data:
        supported = node_id in BOTTOM_NODES
        axis.scatter(x, z, y, color="#dc2626" if supported else "#ff5a5f", s=48, depthshade=False)
        axis.text(x + 0.08, z, y + 0.18, f"N{node_id}\nDOF {dof_numbers(node_id)['UX']}-{dof_numbers(node_id)['RZ']}", fontsize=7)
        if supported:
            axis.scatter(x, z, y - 0.25, marker="^", color="#4b5563", s=125, depthshade=False)
    global_origin = (-0.65, -0.65, -0.65)
    draw_arrow(axis, global_origin, (1.45, 0, 0), "#d97706", "GLOBAL X", fontsize=9, linewidth=2.5)
    draw_arrow(axis, global_origin, (0, 0, 1.45), "#2563eb", "GLOBAL Y", fontsize=9, linewidth=2.5)
    draw_arrow(axis, global_origin, (0, 1.45, 0), "#92400e", "GLOBAL Z", fontsize=9, linewidth=2.5)
    axis.text(global_origin[0] - 0.08, global_origin[1] - 0.12, global_origin[2], "GLOBAL ORIGIN",
              color="#111827", fontsize=7, weight="bold", ha="right")
    axis.set_xlabel("X (m) - lateral")
    axis.set_ylabel("Z (m) - lateral")
    axis.set_zlabel("Y (m) - vertical")
    axis.set_xlim(-1, 7)
    axis.set_ylim(-1, 7)
    axis.set_zlim(-1, 7)
    figure.suptitle("6m x 6m x 6m Cube - Structural Model, Rev. 1\nSupports, member local x/y/z axes, global X/Y/Z axes, beta angles, and MZ releases",
                     fontsize=13, weight="bold", y=0.98)
    axis.legend(handles=[
        plt.Line2D([0], [0], color="#1554d1", lw=2, label="Beam"),
        plt.Line2D([0], [0], color="#14866d", lw=2, label="Column"),
        plt.Line2D([0], [0], marker="o", color="w", markerfacecolor="#dc2626", label="Supported node (pinned)"),
        plt.Line2D([0], [0], marker="o", color="k", markerfacecolor="white", label="Pinned member end (MZ released)"),
        plt.Line2D([0], [0], color="#dc2626", lw=1, label="Local x axis (each member)"),
        plt.Line2D([0], [0], color="#16a34a", lw=1, label="Local y axis (each member)"),
        plt.Line2D([0], [0], color="#7c3aed", lw=1, label="Local z axis (each member)"),
    ], loc="upper left", fontsize=8)
    panel = figure.add_axes([0.70, 0.18, 0.27, 0.65])
    panel.axis("off")
    panel.text(0, 1, "MODEL DATA - REV. 1", family="monospace", fontsize=10, va="top", weight="bold")
    panel.text(0, 0.93, "Geometry\n  Cube edge                 6.0 m\n  Nodes                         8\n  Members                     12\n\nGlobal axes\n  X                         lateral\n  Y                    vertical (up)\n  Z                         lateral\n\nSupports\n  Type                       pinned\n  Nodes                  1, 2, 3, 4\n  Restrained             UX, UY, UZ\n  Released                  RX, RY, RZ\n\nNode degrees of freedom\n  DOF per node                  6\n  Total DOF                    48\n  Restrained DOF               12\n  Active DOF                   36\n  Numbering       (node - 1) x 6 + 1..6\n\nMember end releases\n  Pinned members       1, 3, 5, 7\n  Pattern              pinned in x direction\n  Component             MZ, moment about local z\n  Released end DOF                 8\n  Symbol                 hollow circle\n\nBeta angles\n  Base beam                   0 deg\n  Roof beam                   0 deg\n  Column                     90 deg\n\nLocal axes\n  local x   start node to end node\n  local y   in vertical plane, upward\n  local z   completes right-handed system", family="monospace", fontsize=7.5, va="top")
    panel.add_patch(plt.Rectangle((0, 0), 1, 1, fill=False, edgecolor="#38598a", linewidth=1.5, transform=panel.transAxes, clip_on=False))
    figure.savefig(output_path, dpi=180, facecolor="white")
    plt.close(figure)


output_directory = Path(__file__).resolve().parent
diagram_path = output_directory / "cube_6m_structural_diagram_rev1.png"
workbook_path = output_directory / "cube_6m_structure_rev1.xlsx"
create_structural_diagram(diagram_path)

# Create Workbook
wb = openpyxl.Workbook()

# Sheet 1: Nodes
ws_nodes = wb.active
ws_nodes.title = "Nodes"
ws_nodes.append(["3D Structural Frame - Node Coordinates and DOF (Rev. 1)"])
ws_nodes.append([])
ws_nodes.append(["Node ID", "X (m)", "Y (m) [Vertical]", "Z (m)", "Description", "UX", "UY", "UZ", "RX", "RY", "RZ"])

for n in nodes_data:
    ws_nodes.append(list(n) + [dof_numbers(n[0])[label] for label in DOF_LABELS])

# Sheet 2: Member Incidences
ws_members = wb.create_sheet(title="Member Incidences")
ws_members.append(["3D Structural Frame - Member Incidences (Rev. 1)"])
ws_members.append([])
ws_members.append(["Member ID", "Member Name", "Start Node (i)", "End Node (j)", "Length (m)", "Description", "Calculated Length (m)", "Beta (deg)", "Local x", "Local y", "Local z"])

for mem in members_data_rev1:
    i_val, j_val = mem[2], mem[3]
    r_i = i_val + 3
    r_j = j_val + 3
    formula = f"=SQRT((Nodes!B{r_j}-Nodes!B{r_i})^2 + (Nodes!C{r_j}-Nodes!C{r_i})^2 + (Nodes!D{r_j}-Nodes!D{r_i})^2)"
    start = nodes_data[mem[2] - 1][1:4]
    end = nodes_data[mem[3] - 1][1:4]
    beta = 90.0 if mem[1] in {"M9", "M10", "M11", "M12"} else 0.0
    length, local_x, local_y, local_z = member_axes(start, end, beta)
    ws_members.append([mem[0], mem[1], mem[2], mem[3], mem[4], mem[5], formula, beta,
                       format_vector(local_x), format_vector(local_y), format_vector(local_z)])

ws_dof = wb.create_sheet("Node DOF")
ws_dof.append(["Global DOF numbering - six DOF per node"])
ws_dof.append(["Node", "UX", "UY", "UZ", "RX", "RY", "RZ"])
for node in nodes_data:
    numbers = dof_numbers(node[0])
    ws_dof.append([node[0]] + [numbers[label] for label in DOF_LABELS])

ws_supports = wb.create_sheet("Supports")
ws_supports.append(["Pinned supports at bottom nodes"])
ws_supports.append(["Node", "Support Type", "UX", "UY", "UZ", "RX", "RY", "RZ", "Restrained DOF"])
for node_id in sorted(BOTTOM_NODES):
    numbers = dof_numbers(node_id)
    ws_supports.append([node_id, "Pinned", "Yes", "Yes", "Yes", "No", "No", "No",
                        ", ".join(str(numbers[label]) for label in ("UX", "UY", "UZ"))])

ws_releases = wb.create_sheet("Member Releases")
ws_releases.append(["Beam pinned releases - local MZ at both ends"])
ws_releases.append(["Member", "Start MZ", "End MZ", "Pinned in X direction", "Pinned symbol"])
for mem in members_data_rev1:
    pinned = mem[1] in PINNED_MEMBERS
    ws_releases.append([mem[1], "Yes" if pinned else "No", "Yes" if pinned else "No", "Yes" if pinned else "No", "Hollow circle" if pinned else "No"])

ws_axes = wb.create_sheet("Local Axes")
ws_axes.append(["Member local axes and beta rotation"])
ws_axes.append(["Member", "Beta (deg)", "Local x in global", "Local y in global", "Local z in global"])
for mem in members_data_rev1:
    beta = 90.0 if mem[1] in {"M9", "M10", "M11", "M12"} else 0.0
    _, local_x, local_y, local_z = member_axes(nodes_data[mem[2] - 1][1:4], nodes_data[mem[3] - 1][1:4], beta)
    ws_axes.append([mem[1], beta, format_vector(local_x), format_vector(local_y), format_vector(local_z)])

ws_global = wb.create_sheet("Global Stiffness")
ws_global.append(["Global DOF and release assembly summary"])
ws_global.append(["Member", "Length (m)", "Global translational DOFs", "Released local moment", "Status"])
for mem in members_data_rev1:
    start_dof = dof_numbers(mem[2])
    end_dof = dof_numbers(mem[3])
    length, _, _, _ = member_axes(nodes_data[mem[2] - 1][1:4], nodes_data[mem[3] - 1][1:4], 0.0)
    vector = [start_dof[label] for label in ("UX", "UY", "UZ")] + [end_dof[label] for label in ("UX", "UY", "UZ")]
    ws_global.append([mem[1], length, str(vector), "MZ at both ends" if mem[1] in PINNED_MEMBERS else "None", "Assembled"])

ws_diagram = wb.create_sheet("Structural Diagram")
ws_diagram.append(["Rev. 1 diagram - pinned supports, beta angles, local/global axes, and MZ releases"])
ws_diagram.add_image(ExcelImage(str(diagram_path)), "A3")

for worksheet in wb.worksheets:
    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 24
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="164E63")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(workbook_path)
print(f"[OK] Workbook created: {workbook_path}")
print(f"[OK] Structural diagram created: {diagram_path}")
