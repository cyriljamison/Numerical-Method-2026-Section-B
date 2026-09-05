"""
cube_rev2.py - 6m x 6m x 6m Cube Generator v2 (STAAD/RISA-like extensions)
- Nodes + Members (M1-M12)
- Supports: bottom nodes 1-4 PINNED (Tx,Ty,Tz fixed)
- Beta angle per member (deg, rotation about local x)
- Local axes (x along member, y/z via beta) + Global axes (X,Z lateral, Y vertical)
- DOF: 6 per node (Tx,Ty,Tz,Rx,Ry,Rz) => 48 total
- Member end releases: pinned if Mz released at start/end (show symbol)
- Excel: cube_rev2.xlsx with 6 tabs, enhanced Member Incidences
- Plot: supports, local axes, pinned symbols, global axes, DOF labels

Axis convention:
    X = global lateral
    Z = global lateral
    Y = global vertical
    Node 1 at (0,0,0)
"""

import os
import math
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d.art3d import Line3D, Poly3DCollection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SIZE = 6.0

# ---------------- Nodes ----------------
nodes = {
    1: (0.0, 0.0, 0.0),
    2: (SIZE, 0.0, 0.0),
    3: (SIZE, 0.0, SIZE),
    4: (0.0, 0.0, SIZE),
    5: (0.0, SIZE, 0.0),
    6: (SIZE, SIZE, 0.0),
    7: (SIZE, SIZE, SIZE),
    8: (0.0, SIZE, SIZE),
}

# ---------------- Members ----------------
members = [
    ("M1", 1, 2),
    ("M2", 2, 3),
    ("M3", 3, 4),
    ("M4", 4, 1),
    ("M5", 5, 6),
    ("M6", 6, 7),
    ("M7", 7, 8),
    ("M8", 8, 5),
    ("M9", 1, 5),
    ("M10", 2, 6),
    ("M11", 3, 7),
    ("M12", 4, 8),
]

member_descriptions = {
    "M1": "Bottom front edge (Y=0, parallel to X)",
    "M2": "Bottom right edge (Y=0, parallel to Z)",
    "M3": "Bottom back edge (Y=0, parallel to X)",
    "M4": "Bottom left edge (Y=0, parallel to Z)",
    "M5": "Top front edge (Y=6, parallel to X)",
    "M6": "Top right edge (Y=6, parallel to Z)",
    "M7": "Top back edge (Y=6, parallel to X)",
    "M8": "Top left edge (Y=6, parallel to Z)",
    "M9": "Vertical column - Front-Left (X=0, Z=0)",
    "M10": "Vertical column - Front-Right (X=6, Z=0)",
    "M11": "Vertical column - Back-Right (X=6, Z=6)",
    "M12": "Vertical column - Back-Left (X=0, Z=6)",
}

# Beta angle per member (deg) - rotation about local x (like STAAD BETA)
# Default 0 for all; can be edited per member
member_betas = {
    "M1": 0.0, "M2": 0.0, "M3": 0.0, "M4": 0.0,
    "M5": 0.0, "M6": 0.0, "M7": 0.0, "M8": 0.0,
    "M9": 0.0, "M10": 0.0, "M11": 0.0, "M12": 0.0,
}

# Member end releases: pinned releases for moment about local z (Mz)
# Also example: Fx release placeholder for "pinned at x direction"
# Format: {member: {"start": {"Fx":0/1, "Mz":0/1}, "end": {"Fx":0/1,"Mz":0/1}}}
# 1 = released (pinned), 0 = fixed (rigid)
# Demo: M1-M4 bottom beams pinned at both ends for Mz (typical beam), plus Fx=0 (fixed axial)
member_releases = {
    "M1":  {"start": {"Fx": 0, "Mz": 1}, "end": {"Fx": 0, "Mz": 1}},
    "M2":  {"start": {"Fx": 0, "Mz": 1}, "end": {"Fx": 0, "Mz": 1}},
    "M3":  {"start": {"Fx": 0, "Mz": 1}, "end": {"Fx": 0, "Mz": 1}},
    "M4":  {"start": {"Fx": 0, "Mz": 1}, "end": {"Fx": 0, "Mz": 1}},
    "M5":  {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
    "M6":  {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
    "M7":  {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
    "M8":  {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
    "M9":  {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
    "M10": {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
    "M11": {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
    "M12": {"start": {"Fx": 0, "Mz": 0}, "end": {"Fx": 0, "Mz": 0}},
}

# Supports: bottom nodes 1-4 PINNED (first code style, Tx,Ty,Tz restrained)
# PINNED = Tx,Ty,Tz restrained (0), Rx,Ry,Rz free (1) - pyramid support symbol
# STAAD: SUPPORTS 1 2 3 4 PINNED
supports = {
    1: {"type": "PINNED", "Tx": 0, "Ty": 0, "Tz": 0, "Rx": 1, "Ry": 1, "Rz": 1},
    2: {"type": "PINNED", "Tx": 0, "Ty": 0, "Tz": 0, "Rx": 1, "Ry": 1, "Rz": 1},
    3: {"type": "PINNED", "Tx": 0, "Ty": 0, "Tz": 0, "Rx": 1, "Ry": 1, "Rz": 1},
    4: {"type": "PINNED", "Tx": 0, "Ty": 0, "Tz": 0, "Rx": 1, "Ry": 1, "Rz": 1},
    # Nodes 5-8 free
    5: {"type": "FREE", "Tx": 1, "Ty": 1, "Tz": 1, "Rx": 1, "Ry": 1, "Rz": 1},
    6: {"type": "FREE", "Tx": 1, "Ty": 1, "Tz": 1, "Rx": 1, "Ry": 1, "Rz": 1},
    7: {"type": "FREE", "Tx": 1, "Ty": 1, "Tz": 1, "Rx": 1, "Ry": 1, "Rz": 1},
    8: {"type": "FREE", "Tx": 1, "Ty": 1, "Tz": 1, "Rx": 1, "Ry": 1, "Rz": 1},
}

# DOF mapping: 6 per node -> 48 total
# Order: Tx, Ty, Tz, Rx, Ry, Rz
DOF_LABELS = ["Tx", "Ty", "Tz", "Rx", "Ry", "Rz"]
node_dofs = {nid: [(nid-1)*6 + d +1 for d in range(6)] for nid in nodes}


def get_member_length(i, j):
    x1, y1, z1 = nodes[i]
    x2, y2, z2 = nodes[j]
    return math.sqrt((x2-x1)**2 + (y2-y1)**2 + (z2-z1)**2)

def get_member_orientation(i, j):
    x1, y1, z1 = nodes[i]
    x2, y2, z2 = nodes[j]
    dx, dy, dz = abs(x2-x1), abs(y2-y1), abs(z2-z1)
    eps=1e-9
    if dy>eps and dx<eps and dz<eps: return "Y - Vertical"
    if dx>eps and dy<eps and dz<eps: return "X - Lateral"
    if dz>eps and dx<eps and dy<eps: return "Z - Lateral"
    return "Diagonal / Skewed"

def get_member_beta(mid):
    return member_betas.get(mid, 0.0)

def compute_local_axes(i, j, beta_deg=0.0):
    """Compute local axes x,y,z (unit vectors) for member i->j with beta rotation about x.
    Global: X lateral, Y vertical, Z lateral.
    Returns dict with x,y,z vectors (tuples) and length.
    STAAD-like: x along member, y/z via beta.
    """
    x1,y1,z1 = nodes[i]
    x2,y2,z2 = nodes[j]
    vec = np.array([x2-x1, y2-y1, z2-z1], dtype=float)
    L = np.linalg.norm(vec)
    if L < 1e-12:
        raise ValueError("Zero length member")
    x_vec = vec / L
    # Reference global axes
    global_y = np.array([0.0, 1.0, 0.0])
    global_z = np.array([0.0, 0.0, 1.0])
    # Choose reference for y0/z0
    # If x not parallel to global_y, use global_y as reference
    if abs(abs(np.dot(x_vec, global_y)) - 1.0) < 1e-6:
        # x parallel to Y (vertical column): use global_z as reference
        # For vertical, local y should align with -global X initially (choose)
        global_x = np.array([1.0, 0.0, 0.0])
        y0 = np.cross(global_z, x_vec)
        # If y0 zero, fallback
        if np.linalg.norm(y0) < 1e-9:
            y0 = np.cross(x_vec, global_x)
        y0 = y0 / np.linalg.norm(y0)
        z0 = np.cross(x_vec, y0)
        z0 = z0 / np.linalg.norm(z0)
    else:
        z0 = np.cross(x_vec, global_y)
        z0 = z0 / np.linalg.norm(z0)
        y0 = np.cross(z0, x_vec)
        y0 = y0 / np.linalg.norm(y0)
    # Beta rotation about x
    beta = math.radians(beta_deg)
    cb, sb = math.cos(beta), math.sin(beta)
    y_vec = y0 * cb + z0 * sb
    z_vec = -y0 * sb + z0 * cb
    y_vec = y_vec / np.linalg.norm(y_vec)
    z_vec = z_vec / np.linalg.norm(z_vec)
    return {"x": tuple(x_vec), "y": tuple(y_vec), "z": tuple(z_vec), "L": L, "beta": beta_deg}

def is_member_pinned(mid):
    rel = member_releases.get(mid, {"start":{"Mz":0},"end":{"Mz":0}})
    return rel["start"].get("Mz",0)==1 or rel["end"].get("Mz",0)==1 or rel["start"].get("Fx",0)==1 or rel["end"].get("Fx",0)==1

def get_release_string(mid, end="start"):
    rel = member_releases.get(mid, {"start":{},"end":{}})
    d = rel.get(end, {})
    parts=[]
    if d.get("Fx",0)==1: parts.append("Fx")
    if d.get("Mz",0)==1: parts.append("Mz")
    return ",".join(parts) if parts else "Fixed"

# ---------------- Excel ----------------
def create_excel_file(filename="cube_rev2.xlsx"):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(script_dir, filename) if not os.path.isabs(filename) else filename
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"), top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"))
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # Colors for support (hinge)
    pinned_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    hinge_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Sheet 1: Nodes with Supports and DOF
    ws_nodes = wb.active
    ws_nodes.title = "Nodes"
    ws_nodes.append(["Node","X (m)","Y (m)","Z (m)","Support","Tx","Ty","Tz","Rx","Ry","Rz","DOF Tx","DOF Ty","DOF Tz","DOF Rx","DOF Ry","DOF Rz"])
    ws_nodes.row_dimensions[1].height = 30
    for nid in sorted(nodes):
        x,y,z = nodes[nid]
        sup = supports.get(nid, {"type":"FREE","Tx":1,"Ty":1,"Tz":1,"Rx":1,"Ry":1,"Rz":1})
        dofs = node_dofs[nid]
        ws_nodes.append([nid, x, y, z, sup["type"], sup["Tx"], sup["Ty"], sup["Tz"], sup["Rx"], sup["Ry"], sup["Rz"]] + dofs)
    for col in range(1, 18):
        c = ws_nodes.cell(row=1, column=col); c.font=header_font; c.fill=header_fill; c.alignment=header_align; c.border=thin_border
    for row in range(2, ws_nodes.max_row+1):
        for col in range(1, 18):
            c = ws_nodes.cell(row=row, column=col); c.border=thin_border
            if col in (1,6,7,8,9,10,11,12,13,14,15,16,17): c.alignment=center
            elif col in (2,3,4): c.alignment=center; c.number_format='0.00'
            else: c.alignment=center
            # highlight hinge rows
            if ws_nodes.cell(row=row, column=5).value=="HINGE":
                c.fill=hinge_fill
            elif ws_nodes.cell(row=row, column=5).value=="PINNED":
                c.fill=pinned_fill
        # number formats
        for c_idx in (2,3,4): ws_nodes.cell(row=row, column=c_idx).number_format='0.00'
    widths = [8,10,10,10,10,6,6,6,6,6,6,8,8,8,8,8,8]
    for i,w in enumerate(widths,1): ws_nodes.column_dimensions[chr(64+i) if i<=26 else f"A{i}"].width=w
    # Fix column letters beyond Z
    ws_nodes.column_dimensions["A"].width=8; ws_nodes.column_dimensions["B"].width=10; ws_nodes.column_dimensions["C"].width=10; ws_nodes.column_dimensions["D"].width=10; ws_nodes.column_dimensions["E"].width=10
    ws_nodes.column_dimensions["F"].width=6; ws_nodes.column_dimensions["G"].width=6; ws_nodes.column_dimensions["H"].width=6; ws_nodes.column_dimensions["I"].width=6; ws_nodes.column_dimensions["J"].width=6; ws_nodes.column_dimensions["K"].width=6
    ws_nodes.column_dimensions["L"].width=8; ws_nodes.column_dimensions["M"].width=8; ws_nodes.column_dimensions["N"].width=8; ws_nodes.column_dimensions["O"].width=8; ws_nodes.column_dimensions["P"].width=8; ws_nodes.column_dimensions["Q"].width=8
    ws_nodes.freeze_panes="A2"; ws_nodes.auto_filter.ref=ws_nodes.dimensions
    ws_nodes.sheet_properties.pageSetUpPr.fitToPage=True

    # Sheet 2: Member Incidences enhanced with HINGE support info
    ws_mem = wb.create_sheet("Member Incidences")
    ws_mem.append(["Member","Node i","Node j","Length (m)","Orientation","Beta (deg)","Local x (X,Y,Z)","Local y (X,Y,Z)","Local z (X,Y,Z)","Release i","Release j","Support at i","Support at j","Pinned?","Description"])
    ws_mem.row_dimensions[1].height=30
    for mid,i,j in members:
        length=get_member_length(i,j); orient=get_member_orientation(i,j); beta=get_member_beta(mid)
        axes=compute_local_axes(i,j,beta)
        lx = f"({axes['x'][0]:.2f},{axes['x'][1]:.2f},{axes['x'][2]:.2f})"
        ly = f"({axes['y'][0]:.2f},{axes['y'][1]:.2f},{axes['y'][2]:.2f})"
        lz = f"({axes['z'][0]:.2f},{axes['z'][1]:.2f},{axes['z'][2]:.2f})"
        ri=get_release_string(mid,"start"); rj=get_release_string(mid,"end")
        hinge_i = "YES" if supports.get(i, {}).get("type") in ("HINGE","PINNED") else "NO"
        hinge_j = "YES" if supports.get(j, {}).get("type") in ("HINGE","PINNED") else "NO"
        pinned="YES" if is_member_pinned(mid) else "NO"
        desc=member_descriptions.get(mid,"")
        ws_mem.append([mid,i,j,round(length,2),orient,beta,lx,ly,lz,ri,rj,hinge_i,hinge_j,pinned,desc])
    for col in range(1,16):
        c=ws_mem.cell(row=1, column=col); c.font=header_font; c.fill=header_fill; c.alignment=header_align; c.border=thin_border
    for row in range(2, ws_mem.max_row+1):
        for col in range(1,16):
            c=ws_mem.cell(row=row, column=col); c.border=thin_border
            if col in (15,7,8,9): c.alignment=left_wrap
            else: c.alignment=center
            if col==4: c.number_format='0.00'
            if col==6: c.number_format='0.0'
            if ws_mem.cell(row=row, column=14).value=="YES":
                c.fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            # highlight hinge connections
            if col in (12,13) and c.value=="YES":
                c.fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    widths2=[10,8,8,10,14,10,16,16,16,12,12,10,10,10,42]
    cols=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O"]
    for col,w in zip(cols,widths2): ws_mem.column_dimensions[col].width=w
    ws_mem.freeze_panes="A2"; ws_mem.auto_filter.ref=ws_mem.dimensions
    ws_mem.sheet_properties.pageSetUpPr.fitToPage=True

    # Sheet 3: Supports
    ws_sup = wb.create_sheet("Supports")
    ws_sup.append(["Node","X","Y","Z","Type","Tx","Ty","Tz","Rx","Ry","Rz","Comments"])
    ws_sup.row_dimensions[1].height=22
    for nid in sorted(nodes):
        if supports[nid]["type"]!="FREE":
            x,y,z=nodes[nid]; sup=supports[nid]
            if sup["type"]=="HINGE":
                comments="Hinged: translations restrained, rotations free (hinge support)"
            elif sup["type"]=="PINNED":
                comments="Pinned: translations restrained, rotations free"
            else:
                comments=""
            ws_sup.append([nid,x,y,z,sup["type"],sup["Tx"],sup["Ty"],sup["Tz"],sup["Rx"],sup["Ry"],sup["Rz"],comments])
    for col in range(1,13):
        c=ws_sup.cell(row=1, column=col); c.font=header_font; c.fill=header_fill; c.alignment=header_align; c.border=thin_border
    for row in range(2, ws_sup.max_row+1):
        for col in range(1,13):
            c=ws_sup.cell(row=row, column=col); c.border=thin_border; c.alignment=center if col!=12 else left_wrap
            if col in (2,3,4): c.number_format='0.00'
    w_sup=[8,10,10,10,10,6,6,6,6,6,6,42]
    for col,w in zip(["A","B","C","D","E","F","G","H","I","J","K","L"],w_sup): ws_sup.column_dimensions[col].width=w
    ws_sup.freeze_panes="A2"; ws_sup.auto_filter.ref=ws_sup.dimensions

    # Sheet 4: DOF Map
    ws_dof = wb.create_sheet("DOF Map")
    ws_dof.append(["Node","DOF Tx","DOF Ty","DOF Tz","DOF Rx","DOF Ry","DOF Rz","Status Tx","Status Ty","Status Tz","Status Rx","Status Ry","Status Rz"])
    ws_dof.row_dimensions[1].height=30
    for nid in sorted(nodes):
        dofs=node_dofs[nid]; sup=supports[nid]
        statuses=[]
        for k in DOF_LABELS:
            statuses.append("Restrained" if sup[k]==0 else "Free")
        ws_dof.append([nid]+dofs+statuses)
    for col in range(1,14):
        c=ws_dof.cell(row=1, column=col); c.font=header_font; c.fill=header_fill; c.alignment=header_align; c.border=thin_border
    for row in range(2, ws_dof.max_row+1):
        for col in range(1,14):
            c=ws_dof.cell(row=row, column=col); c.border=thin_border; c.alignment=center
            # color restrained
            if col>=8 and c.value=="Restrained":
                c.fill=PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
            elif col>=8 and c.value=="Free":
                c.fill=PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M"]:
        ws_dof.column_dimensions[col].width=10
    ws_dof.freeze_panes="A2"; ws_dof.auto_filter.ref=ws_dof.dimensions

    # Sheet 5: Local Axes
    ws_local = wb.create_sheet("Local Axes")
    ws_local.append(["Member","i","j","Beta (deg)","Local x (X,Y,Z)","Local y (X,Y,Z)","Local z (X,Y,Z)","Length (m)","Global Axes Note"])
    ws_local.row_dimensions[1].height=30
    for mid,i,j in members:
        beta=get_member_beta(mid); axes=compute_local_axes(i,j,beta)
        lx = f"({axes['x'][0]:.2f},{axes['x'][1]:.2f},{axes['x'][2]:.2f})"
        ly = f"({axes['y'][0]:.2f},{axes['y'][1]:.2f},{axes['y'][2]:.2f})"
        lz = f"({axes['z'][0]:.2f},{axes['z'][1]:.2f},{axes['z'][2]:.2f})"
        ws_local.append([mid,i,j,beta, lx, ly, lz, round(axes['L'],2), "Global: X,Z lateral, Y vertical"])
    for col in range(1,10):
        c=ws_local.cell(row=1, column=col); c.font=header_font; c.fill=header_fill; c.alignment=header_align; c.border=thin_border
    for row in range(2, ws_local.max_row+1):
        for col in range(1,10):
            c=ws_local.cell(row=row, column=col); c.border=thin_border; c.alignment=center if col not in (5,6,7) else left_wrap
            if col==8: c.number_format='0.00'
            if col==4: c.number_format='0.0'
    for col,w in zip(["A","B","C","D","E","F","G","H","I"],[10,8,8,10,20,20,20,10,28]): ws_local.column_dimensions[col].width=w
    ws_local.freeze_panes="A2"; ws_local.auto_filter.ref=ws_local.dimensions

    # Sheet 6: Info/Summary
    ws_info = wb.create_sheet("Info")
    ws_info.append(["Item","Value / Description"])
    ws_info.append(["Cube size","6m x 6m x 6m"])
    ws_info.append(["Nodes",8])
    ws_info.append(["Members",12])
    ws_info.append(["Supports","Nodes 1-4 PINNED (Tx,Ty,Tz=0, Rx,Ry,Rz=1) - pyramid support (first code)"])
    ws_info.append(["DOF per node",6])
    ws_info.append(["Total DOF",48])
    ws_info.append(["Beta default","0 deg for all members (M1-M12)"])
    ws_info.append(["Pinned beams","M1-M4 bottom beams: Mz released at both ends (Fx fixed) -> pinned symbol shown"])
    ws_info.append(["Global axes","X (lateral), Y (vertical), Z (lateral) - triad on left side at (-1.7,0.4,3)"])
    ws_info.append(["Support note","Bottom nodes 1-4 shown as PINNED pyramid (first code), also reflected in Member Incidences Support at i/j"])
    ws_info.append(["Local axes","x along member, y/z via beta rotation about x"])
    ws_info.append(["Files","cube_rev2.py, cube_rev2.xlsx, cube_rev2.png"])
    for col in range(1,3):
        c=ws_info.cell(row=1, column=col); c.font=header_font; c.fill=header_fill; c.alignment=header_align; c.border=thin_border
    for row in range(2, ws_info.max_row+1):
        for col in range(1,3):
            c=ws_info.cell(row=row, column=col); c.border=thin_border; c.alignment=left_wrap if col==2 else center
    ws_info.column_dimensions["A"].width=18; ws_info.column_dimensions["B"].width=65
    ws_info.sheet_properties.pageSetUpPr.fitToPage=True

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage=True
    wb.save(filepath)
    print(f"Excel file created: {filepath} with {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    return filepath

# ---------------- Plot ----------------
def plot_cube(save_path=None, show=True, show_local_axes=True, show_dof=True):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if save_path is None: save_path=os.path.join(script_dir, "cube_rev2.png")
    elif not os.path.isabs(save_path): save_path=os.path.join(script_dir, save_path)
    fig = plt.figure(figsize=(11,9))
    ax = fig.add_subplot(111, projection="3d")
    xs=[c[0] for c in nodes.values()]; ys=[c[1] for c in nodes.values()]; zs=[c[2] for c in nodes.values()]
    ax.scatter(xs, ys, zs, color="red", s=90, depthshade=False, label="Nodes", zorder=5, edgecolors="black", linewidths=0.5)
    # Node labels with DOF
    for nid,(x,y,z) in nodes.items():
        dofs=node_dofs[nid]
        label=f"{nid}"
        if show_dof:
            # small DOF label below node
            dof_str = f"({dofs[0]}-{dofs[5]})"
            ax.text(x+0.12, y+0.12, z+0.12, label, fontsize=11, color="darkred", weight="bold")
            ax.text(x+0.12, y-0.35, z+0.12, dof_str, fontsize=6, color="dimgray")
        else:
            ax.text(x+0.15, y+0.15, z+0.15, label, fontsize=11, color="darkred", weight="bold")
        # Support symbol - PINNED pyramid (first code) for PINNED
        if supports[nid]["type"]=="PINNED":
            s=0.9
            base_y = y - 0.7
            corners = [(x-s/2, base_y, z-s/2), (x+s/2, base_y, z-s/2), (x+s/2, base_y, z+s/2), (x-s/2, base_y, z+s/2)]
            for cx,cy,cz in corners:
                ax.plot([x,cx],[y,cy],[z,cz], color="black", lw=1.2)
            for k in range(4):
                x1,y1,z1=corners[k]; x2,y2,z2=corners[(k+1)%4]
                ax.plot([x1,x2],[y1,y2],[z1,z2], color="black", lw=1.5)
            ax.text(x, base_y-0.15, z, "PINNED", fontsize=6, color="black", ha="center", weight="bold")
    # Global axes on left side of cube (classic)
    g_len=1.3
    gx, gy, gz = -1.7, 0.4, SIZE/2  # left side, centered in Z
    ax.quiver(gx,gy,gz, g_len,0,0, color="red", linewidth=2.2, arrow_length_ratio=0.14)
    ax.quiver(gx,gy,gz, 0,g_len,0, color="green", linewidth=2.2, arrow_length_ratio=0.14)
    ax.quiver(gx,gy,gz, 0,0,g_len, color="blue", linewidth=2.2, arrow_length_ratio=0.14)
    ax.text(gx+g_len+0.15, gy, gz, "X (lateral)", color="red", fontsize=8, weight="bold")
    ax.text(gx, gy+g_len+0.15, gz, "Y (vertical)", color="green", fontsize=8, weight="bold")
    ax.text(gx, gy, gz+g_len+0.15, "Z (lateral)", color="blue", fontsize=8, weight="bold")
    # small origin marker for global axes
    ax.scatter([gx],[gy],[gz], s=40, color="black", zorder=6)
    # Members
    for mid,i,j in members:
        x1,y1,z1=nodes[i]; x2,y2,z2=nodes[j]
        color="steelblue"; lw=2.4
        # Highlight pinned members lighter?
        if is_member_pinned(mid):
            color="#1f77b4"; lw=2.6
        ax.add_line(Line3D([x1,x2],[y1,y2],[z1,z2], color=color, lw=lw))
        mx,my,mz=(x1+x2)/2,(y1+y2)/2,(z1+z2)/2
        ax.text(mx,my,mz,f"{mid}", fontsize=7, color="darkgreen", alpha=0.9, ha="center", va="center",
                bbox=dict(boxstyle="round,pad=0.15", fc="white", ec="green", alpha=0.6))
        # Beta label
        beta=get_member_beta(mid)
        # Local axes at midpoint
        if show_local_axes:
            axes=compute_local_axes(i,j,beta)
            scale=0.9
            # x red, y green, z blue (local)
            ax.quiver(mx,my,mz, axes["x"][0]*scale, axes["x"][1]*scale, axes["x"][2]*scale, color="red", linewidth=1.2, arrow_length_ratio=0.2)
            ax.quiver(mx,my,mz, axes["y"][0]*scale, axes["y"][1]*scale, axes["y"][2]*scale, color="green", linewidth=1.2, arrow_length_ratio=0.2)
            ax.quiver(mx,my,mz, axes["z"][0]*scale, axes["z"][1]*scale, axes["z"][2]*scale, color="blue", linewidth=1.2, arrow_length_ratio=0.2)
            # beta text slightly offset
            # ax.text(mx+0.1,my+0.1,mz+0.1, f"β={beta:.0f}°", fontsize=5, color="purple")
        # Pinned symbols at ends if Mz or Fx released
        rel=member_releases.get(mid, {"start":{},"end":{}})
        for end, (xi, yi, zi) in [("start",(x1,y1,z1)), ("end",(x2,y2,z2))]:
            d=rel.get(end,{})
            if d.get("Mz",0)==1 or d.get("Fx",0)==1:
                # draw small circle/hollow marker at end offset slightly towards mid
                # offset 0.12*L towards center to avoid node overlap
                ox, oy, oz = mx, my, mz
                # position 8% from node towards center
                px = xi*0.92 + mx*0.08
                py = yi*0.92 + my*0.08
                pz = zi*0.92 + mz*0.08
                ax.scatter([px],[py],[pz], s=90, facecolors="white", edgecolors="orange", linewidths=1.8, marker="o", zorder=6)
                # small text
                ax.text(px,py+0.18,pz, "pin", fontsize=5, color="orange", ha="center", weight="bold")

    ax.set_xlabel("X (m) - Lateral (Global)")
    ax.set_ylabel("Y (m) - Vertical (Global)")
    ax.set_zlabel("Z (m) - Lateral (Global)")
    ax.set_xlim(-2.2, SIZE+1); ax.set_ylim(-1.5, SIZE+1); ax.set_zlim(-1, SIZE+1)
    ax.set_box_aspect((1,1,1))
    ax.set_title("6m Cube v2 - STAAD/RISA-like:\nPINNED Supports (1-4 pyramid, first code), Beta=0°, Local axes (R:x G:y B:z), Pinned M1-M4 (Mz), DOF 1-48, Global axis on left", fontsize=10)
    ax.view_init(elev=22, azim=-35)
    ax.grid(True, alpha=0.3)
    # Custom legend handles
    import matplotlib.patches as mpatches
    from matplotlib.lines import Line2D
    legend_elems=[
        Line2D([0],[0], marker='o', color='w', markerfacecolor='red', markeredgecolor='black', markersize=8, label='Nodes (1-8)'),
        Line2D([0],[0], color='steelblue', lw=2, label='Members M1-M12'),
        Line2D([0],[0], marker='o', color='w', markerfacecolor='white', markeredgecolor='orange', markersize=8, label='Pinned end (Mz/Fx)'),
        Line2D([0],[0], color='black', lw=1.5, label='Pinned support (1-4) pyramid'),
        Line2D([0],[0], color='red', lw=2, label='Global/Local X (left axis)'),
        Line2D([0],[0], color='green', lw=2, label='Global/Local Y'),
        Line2D([0],[0], color='blue', lw=2, label='Global/Local Z'),
    ]
    ax.legend(handles=legend_elems, loc="upper left", bbox_to_anchor=(0.02,0.98), fontsize=7, framealpha=0.9)
    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    print(f"Plot saved: {save_path} (local_axes={show_local_axes}, dof={show_dof})")
    if show:
        plt.show()
    else:
        plt.close(fig)
    return save_path

def main():
    print("Creating 6m Cube v2 - STAAD/RISA-like...")
    print(f"SIZE={SIZE} m, Nodes={len(nodes)}, Members={len(members)}, DOF={len(nodes)*6}")
    print("\nNodes with Supports & DOF (Y vertical):")
    print(f"{'Node':<6}{'X':<6}{'Y':<6}{'Z':<6}{'Support':<8}{'Tx':<4}{'Ty':<4}{'Tz':<4}{'DOF'}")
    for nid in sorted(nodes):
        x,y,z=nodes[nid]; sup=supports[nid]; dofs=node_dofs[nid]
        print(f"{nid:<6}{x:<6.1f}{y:<6.1f}{z:<6.1f}{sup['type']:<8}{sup['Tx']:<4}{sup['Ty']:<4}{sup['Tz']:<4}{dofs}")
    print("\nMembers with Beta, Local Axes, Releases:")
    print(f"{'Member':<8}{'i':<4}{'j':<4}{'L':<6}{'Beta':<6}{'Local x':<16}{'Local y':<16}{'Local z':<16}{'Rel i':<10}{'Rel j':<10}{'Pinned'}")
    for mid,i,j in members:
        ax=compute_local_axes(i,j,get_member_beta(mid))
        print(f"{mid:<8}{i:<4}{j:<4}{ax['L']:<6.2f}{ax['beta']:<6.1f}{str(tuple(round(v,2) for v in ax['x'])):<16}{str(tuple(round(v,2) for v in ax['y'])):<16}{str(tuple(round(v,2) for v in ax['z'])):<16}{get_release_string(mid,'start'):<10}{get_release_string(mid,'end'):<10}{is_member_pinned(mid)}")
    excel_path=create_excel_file("cube_rev2.xlsx")
    plot_path=plot_cube(save_path="cube_rev2.png", show=True, show_local_axes=True, show_dof=True)
    print("\nDone v2.")
    print(f" - Excel: {excel_path}")
    print(f" - Plot : {plot_path}")
    print("Sheets: Nodes, Member Incidences, Supports, DOF Map, Local Axes, Info")

if __name__=="__main__":
    main()
